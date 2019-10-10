#!/usr/bin/python
# -*- coding: UTF-8 
# author: Ian
# Please,you must believe yourself who can do it beautifully !
"""
Are you OK?
https://www.jd.com/
爬取京东医药
"""

from bs4 import BeautifulSoup
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import os
from openpyxl import Workbook,load_workbook
import logging

BASE_URL = "https://search.jd.com/Search?keyword={0}&enc=utf-8&wq={1}&pvid=e594be63471b41b2a6821999b9c2e65a"
ROOT_PATH = "./file/image"
RESOURCE = r"./resource/商品100.xlsx"

class MedicineSpiderJD(object):
    keys = []
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # 使用无头谷歌浏览器模式
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        # 指定谷歌浏览器路径
        # webdriver.Chrome(chrome_options=chrome_options, executable_path='/usr/local/bin/chromedriver')
        webdriver.Chrome(options=chrome_options)
        self.driver = webdriver.Chrome()

        file_name = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        file = "./resource/{0}.log".format(file_name)
        logging.basicConfig(filename=file, filemode="w", format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                            datefmt="%d-%M-%Y %H:%M:%S", level=logging.DEBUG)

    def requestUrl(self, url):
        self.driver.get(url=url)
        # 等待1s
        time.sleep(1)

        # 获取响应内容
        content = self.driver.page_source.encode('utf-8')
        content = str(content, "utf-8")
        return content



    def close(self):
        self.driver.close()
        self.driver.quit()

    # 解析数据
    def parseHtml(self, data, keys):
        # print(data)
        soup = BeautifulSoup(data, "lxml")

        detail_urls = []
        img_urls = []
        img_list = soup.find_all("div", "p-img")
        item_list = []
        if len(img_list) > 7:
            for i in range(0, 3):
                item_list.append(img_list[i])

        for div in item_list:
            try:
                # title = div.find("a").attrs["title"]
                detail_url = "http:"+ div.find("a").attrs["href"]
                img_url = "http:"+ div.find("img").attrs["src"]
            except Exception as e:
                print("Exception#", e)
                img_url = "http:"+div.find("img").attrs["data-lazy-img"]


            print(detail_url, img_url)

            detail_urls.append(detail_url)
            img_urls.append(img_url)

        print(len(detail_urls))

        index = 0
        # 过入详细页
        for url in detail_urls:
            print("请求详细：",url)
            content = self.requestUrl(url)
            # 找不了，就不在向下执行了
            if self.parseDetailHtml(img_urls[index], content, keys):
                break
            index += 1


    # 解析详细
    def parseDetailHtml(self, img_url, data, keys):
        # print(data)
        soup = BeautifulSoup(data, "html.parser")
        # soup = BeautifulSoup(data, "lxml")
        isFinding = False
        try:
            div = soup.find_all("div", "Ptable-item")
            dts = div[0].contents[3].find_all("dl")
            for dt in dts:
                label = dt.find("dt").get_text()
                name = dt.find("dd").get_text()
                label = label.strip()
                name = name.strip()
                print(label, name, "    -----------------批注文号#"+keys[2])

                # 判断两个ID是否相同
                if label == "批准文号" and name == keys[2]:
                    isFinding = True
                    print("列表图片地址：", img_url)
                    img_url = img_url.replace("/n5/", "/n1/")
                    path = "{0}/{1}".format(ROOT_PATH, keys[0])
                    index = 0
                    name = "{0}".format(index)
                    # name = "{0}_{1}_{2}".format(keys[4], keys[1], index)
                    self.download_image(img_url, path, name)
                    lis = soup.find_all("div", "spec-items")[0].contents[1]
                    imgs = lis.find_all("img")
                    for img in imgs:
                        try:
                            # n0~n5 从800x800~50x50
                            # url = "http://img11.360buyimg.com/n0/" + img.attrs["src"]
                            url = "http://img11.360buyimg.com/n0/"+img.attrs["data-url"]
                        except Exception as e:
                            print("Exception#", e)

                        index += 1
                        name = "{0}".format(index)
                        # name = "{0}_{1}_{2}".format(keys[4], keys[1], index)
                        self.download_image(url, path, name)
                    break;
        except Exception as e:
            print("Exception#", e)
            logging.exception("#### 解析异常 #### " + keys[4])

        return isFinding

    # 下载图片
    def download_image(self, url, path, name):
        if not os.path.exists(path):
            os.makedirs(path)

        file = path + "/" + name + ".jpg"
        print("下载路径：", file, url)
        logging.info("#### 下载路径 #### " + file)
        logging.info("#### 下载URL #### " + url)

        out = open(file, "wb")
        out.write(requests.get(url).content)
        out.close()

    # 解析全部数据
    def spider(self):
        # 获取所有关键字
        self.keys = self.get_resource(RESOURCE)
        for key in self.keys:
            # url_key = key[4] +" "+key[1]
            url_key = key[4]
            url = BASE_URL.format(url_key, url_key)
            try:
                data = self.requestUrl(url)
                self.parseHtml(data, key)
            except Exception as e:
                print(e)
                logging.exception("#### 找不到关键字 #### "+url_key)
                print("spider#### 找不到关键字 ####", url_key)


        spider.close()

    def get_resource(self, resource):
        print(resource)
        # 找到需要xlsx文件的位置
        workBook = load_workbook(resource)

        # 如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
        # sheets = workBook.get_sheet_names()  # 从名称获取sheet
        sheets = workBook.sheetnames  # 从名称获取sheet
        print(sheets)

        medicine_values = []

        # 通过名字打开Sheet
        for i in range(0, len(sheets)):
            # booksheet = workBook.get_sheet_by_name(sheets[i])
            booksheet = workBook[sheets[i]]
            print(booksheet.title)
            # 获取sheet页的行数据
            rows = booksheet.rows
            # 获取sheet页的列数据
            columns = booksheet.columns
            j = 0
            # 迭代所有的行
            for row in rows:
                j = j + 1
                line = [col.value for col in row]
                # print(j, line)

                # 获取每一行的数据
                cell_data_1 = booksheet.cell(row=j, column=1).value  # ERP
                cell_data_2 = booksheet.cell(row=j, column=2).value  # 生产商
                cell_data_3 = booksheet.cell(row=j, column=3).value  # 国家批文
                cell_data_4 = booksheet.cell(row=j, column=4).value  # 药品ID
                cell_data_5 = booksheet.cell(row=j, column=6).value  # 药品简称
                print(cell_data_1, cell_data_2, cell_data_3, cell_data_4, cell_data_5)
                medicine_values.append([cell_data_1, cell_data_2, cell_data_3, cell_data_4, cell_data_5])

        print("获取数据总条目：", len(medicine_values))
        medicine_values.pop(0)
        print(medicine_values)
        return medicine_values


def test_search():
    spider = MedicineSpiderJD()
    key = ["432409","黄石市力康药业有限公司","国药准字Z20026801","1251196",	"Z00095000140010","伤湿止痛膏(OTC)"]
    # url_key = key[5]
    # url_key = url_key.replace("(OTC)", "")
    # url_key = url_key.replace("(处方药)", "")
    # url_key += key[1][0:6]
    url_key = "小金片太极"
    url = BASE_URL.format(url_key, url_key)
    data = spider.requestUrl(url)
    spider.parseHtml(data, key)
    spider.close()

def test_detail():
    spider = MedicineSpiderJD()
    # 1340	天津华津制药有限公司	国药准字H10910053	1110519	X00400400020010	格列齐特片(II)(处方药)
    keys = ["1340", "天津华津制药有限公司", "国药准字H10900089", "1110519", "格列齐特片(II)(处方药)"]
    url = "http://item.jd.com/2947566.html"
    img_url = "http://img11.360buyimg.com/n5/jfs/t5506/319/396551689/270500/a77091d3/58fef1b6N20a71c72.jpg"
    data = spider.requestUrl(url)
    spider.parseDetailHtml(img_url, data, keys)
    spider.close()

def test_spider():
    spider = MedicineSpiderJD()
    spider.spider()
    # spider.get_resource(RESOURCE)
    spider.close()

def compare(name, dirs):
    for file in dirs:
        if name == file:
            return True

        # 没有下载成功的商品，也需要重新下载
        child_dir = ROOT_PATH + "/"+file
        if os.path.isdir(child_dir) and len(os.listdir(child_dir)) < 1:
            return True
    return False

def test_spider_01():
    spider = MedicineSpiderJD()
    # data = spider.get_resource("./resource/medicine.xlsx")
    data = spider.get_resource(RESOURCE)
    dirs = os.listdir(ROOT_PATH)
    for key in data:
        name = str(key[0])
        flag = compare(name, dirs)
        print(key)
        print("比较 {0} 是否已经存在:{1} ".format(name, flag))

        if not flag:
            url_key = key[4]
            # 替换掉多余的关键字
            url_key = url_key.replace("(OTC)", "")
            url_key = url_key.replace("(处方药)", "")
            u_key = "{0} {1}".format(url_key, key[1][2:4])
            url_key = u_key
            print("********************* 关键字：", url_key)
            logging.info("********************* 关键字："+ url_key)
            url = BASE_URL.format(url_key, url_key)
            logging.info(url)
            data = spider.requestUrl(url)
            spider.parseHtml(data, key)

    spider.close()


def save_resource(resource, data):
    print(resource)

    # 找到需要xlsx文件的位置
    workBook = Workbook()
    # 获取当前活跃的sheet,默认是第一个sheet
    bookAction = workBook.active
    bookAction.title = "药品清单"


    for i in range(0, len(data)):
        print(data[i])

        for j in range(0, len(data[i])):
            bookAction.cell(row=i+1, column=j+1).value = data[i][j]

    workBook.save(filename=resource)

def get_all_download_Medicine():
    spider = MedicineSpiderJD()
    data = spider.get_resource(RESOURCE)
    dirs = os.listdir(ROOT_PATH)
    download_medicine = []
    for key in data:
        name = str(key[0])
        flag = compare(name, dirs)
        print(key)
        print("比较 {0} 是否已经存在:{1} ".format(name, flag))
        if not flag:
            download_medicine.append(key)

    download_medicine.insert(0, ["ERP商品ID","ERP生产厂家", "ERP批注文号","药品id",	"药品名称"])
    save_resource("./resource/medicine.xlsx", download_medicine)
    print(len(download_medicine))
    spider.close()

if __name__ == "__main__":
    print("开始执行任务")
    # test_detail()
    # test_spider()
    # test_search()
    # test_spider_01()
    get_all_download_Medicine()

