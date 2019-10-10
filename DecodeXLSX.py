#!/usr/bin/python
# -*- coding: UTF-8 
# author: Ian
# Please,you must believe yourself who can do it beautifully !
"""
Are you OK?

第一步：先把xlsx中所有item拿到
第二步：分步搜索，解析相应的数据
第三步：保存到本地

如果需要拿到商品的详细，那么就需要解析js，跳转详情
"""
import requests
import re
import codecs
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
import time
from datetime import datetime
from urllib.parse import quote,unquote

RESOURCE = r"/Users/ianchang/Downloads/商品及库存.xlsx"
root = "./resource"
if not os.path.exists(root):
    os.makedirs(root)

SAVE_RESOURCE_FILE = root+"/商品100.xlsx"

def save_resource(resource, data):
    print(resource)

    # 找到需要xlsx文件的位置
    workBook = Workbook()
    # 获取当前活跃的sheet,默认是第一个sheet
    bookAction = workBook.active
    bookAction.title = "药品清单"


    for i in range(0, len(data)):
        print(data[i])

        for j in range(0, len(data[i])):
            bookAction.cell(row=i+1, column=j+1).value = data[i][j]

    workBook.save(filename=resource)

def get_resource(resource):
    print(resource)
    # 找到需要xlsx文件的位置
    workBook = load_workbook(resource)

    # 如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
    # sheets = workBook.get_sheet_names()  # 从名称获取sheet
    sheets = workBook.sheetnames  # 从名称获取sheet
    print(sheets)

    medicine_types = ["药品编号ID", "药品名称", "药品简称", "重量", "包装", "类型", "厂家", "日期"]
    medicine_values = []

    # 通过名字打开Sheet
    for i in range(0, len(sheets)):
        # booksheet = workBook.get_sheet_by_name(sheets[i])
        booksheet = workBook[sheets[i]]
        print(booksheet.title)
        # 获取sheet页的行数据
        rows = booksheet.rows
        # 获取sheet页的列数据
        columns = booksheet.columns
        j = 0
        # 迭代所有的行
        for row in rows:
            j = j + 1
            line = [col.value for col in row]
            # print(j, line)

            # 获取每一行的数据
            cell_data_1 = booksheet.cell(row=j, column=1).value  # 药品编号ID
            cell_data_2 = booksheet.cell(row=j, column=6).value  # 药品名称
            cell_data_3 = booksheet.cell(row=j, column=7).value  # 药品简称
            cell_data_4 = booksheet.cell(row=j, column=8).value  # 重量
            cell_data_5 = booksheet.cell(row=j, column=9).value  # 包装
            cell_data_6 = booksheet.cell(row=j, column=10).value  # 类型
            cell_data_7 = booksheet.cell(row=j, column=11).value  # 厂家
            cell_data_8 = booksheet.cell(row=j, column=34).value  # 日期
            print(cell_data_1, cell_data_2, cell_data_3, cell_data_4, cell_data_5, cell_data_6, cell_data_7, cell_data_8)
            medicine_values.append([cell_data_1, cell_data_2, cell_data_3, cell_data_4, cell_data_5, cell_data_6, cell_data_7, cell_data_8])

    print("获取数据总条目：", len(medicine_values))
    medicine_values.pop(0)
    medicine_values.insert(0, medicine_types)
    print(medicine_values)
    return medicine_values


def get_resource_01(resource):
    print(resource)
    # 找到需要xlsx文件的位置
    workBook = load_workbook(resource)

    # 如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
    # sheets = workBook.get_sheet_names()  # 从名称获取sheet
    sheets = workBook.sheetnames  # 从名称获取sheet
    print(sheets)

    medicine_types = ["ERP商品ID", "药品id", "药品名称"]
    medicine_values = []
    sheets.remove("Sheet2")

    # 通过名字打开Sheet
    for i in range(0, len(sheets)):
        # booksheet = workBook.get_sheet_by_name(sheets[i])
        booksheet = workBook[sheets[i]]
        print(booksheet.title)
        # 获取sheet页的行数据
        rows = booksheet.rows
        # 获取sheet页的列数据
        columns = booksheet.columns
        j = 0
        # 迭代所有的行
        for row in rows:
            j = j + 1
            line = [col.value for col in row]
            # print(j, line)

            # 获取每一行的数据
            cell_data_1 = booksheet.cell(row=j, column=1).value  # 药品编号ID
            cell_data_2 = booksheet.cell(row=j, column=2).value  # 药品名称
            cell_data_3 = booksheet.cell(row=j, column=4).value  # 药品简称
            print(cell_data_1, cell_data_2, cell_data_3)
            medicine_values.append([cell_data_1, cell_data_2, cell_data_3])

    print("获取数据总条目：", len(medicine_values))
    medicine_values.pop(0)
    medicine_values.insert(0, medicine_types)
    print(medicine_values)
    return medicine_values

def decodeDocument(doc):
    soup = BeautifulSoup(doc, 'html.parser')
    divs = soup.find_all('div', class_='itemSearchResultCon')
    print(len(divs))
    for i in range(0, len(divs)):
        div = divs[i]
        # print("--------->",div)
        href = div.find("a", attrs={'class': 'product_pic pro_img'})
        detail_url = href.get("href")
        name = div.find("img").get("alt")
        image_url = div.find("img").get("src")
        price = div.find('p', class_='price').find("span").get_text()
        print("详细:{0} 名称：{1} 图片:{2} 价格:{3}".format(detail_url, name, image_url, price))


def download_page(url):
    """获取url地址页面内容"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.80 Safari/537.36'
    }

    response = requests.get(url, headers=headers)
    data = response.text
    print(data)
    return data

def get_html(url, key):
    text = url+key
    print(text)
    return download_page(text)

if __name__ == "__main__":
    print("Hello World")
    # medicines = get_resource(RESOURCE)
    # save_resource(SAVE_RESOURCE_FILE, medicines)

    medicines = get_resource_01("/Users/ianchang/Downloads/医保品种目录1043-科创.xlsx")
    save_resource(SAVE_RESOURCE_FILE, medicines)

    # url = "https://www.111.com.cn/search/search.action?keyWord="
    # content = get_html(url, "布洛芬缓释胶囊(芬必得)")
    # decodeDocument(content)

    # text = "布洛芬缓释胶囊（芬必得）&category=953710&gotoPage=1"
    # data = quote(text, 'utf-8')
    # print(data)
    # data = "%25E5%25B8%2583%25E6%25B4%259B%25E8%258A%25AC%25E7%25BC%2593%25E9%2587%258A%25E8%2583%25B6%25E5%259B%258A(%25E8%258A%25AC%25E5%25BF%2585%25E5%25BE%2597)"
    # data = "%25E8%258A%25AC%25E5%25BF%2585%25E5%25BE%2597"
    # print(unquote(data, 'utf-8'))