#!/usr/bin/python
# -*- coding: UTF-8 
# author: Ian
# Please,you must believe yourself who can do it beautifully !
"""
Are you OK?
http://yao.xywy.com/
医药网  1药网
http://yao.xywy.com/search/?q=%E5%90%89%E6%9E%97%E9%91%AB%E8%BE%89+%E7%89%9B%E9%BB%84%E8%A7%A3%E6%AF%92%E7%89%87&sort=complex&pricefilter=0
"""

from bs4 import BeautifulSoup
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import os
from openpyxl import Workbook, load_workbook
import logging

BASE_URL = "http://yao.xywy.com/search/?q={0}&sort=complex&pricefilter=0"
ROOT_PATH = "./file/image"
RESOURCE = r"./resource/medicine.xlsx"

class MedicineSpider_YAOPIN(object):
    keys = []

    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # 使用无头谷歌浏览器模式
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        # 指定谷歌浏览器路径
        # webdriver.Chrome(chrome_options=chrome_options, executable_path='/usr/local/bin/chromedriver')
        webdriver.Chrome(options=chrome_options)
        self.driver = webdriver.Chrome()

        file_name = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        file = "./resource/{0}.log".format(file_name)
        logging.basicConfig(filename=file, filemode="w", format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                            datefmt="%d-%M-%Y %H:%M:%S", level=logging.DEBUG)

    def close(self):
        self.driver.close()
        self.driver.quit()

    def requestUrl(self, url):
        print("requestUrl#", url)
        self.driver.get(url=url)
        # 等待1s
        time.sleep(1)

        # 获取响应内容
        content = self.driver.page_source.encode('utf-8')
        content = str(content, "utf-8")
        return content

    def get_resource(self, resource):
        print(resource)
        # 找到需要xlsx文件的位置
        workBook = load_workbook(resource)

        # 如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
        # sheets = workBook.get_sheet_names()  # 从名称获取sheet
        sheets = workBook.sheetnames  # 从名称获取sheet
        print(sheets)

        medicine_values = []

        # 通过名字打开Sheet
        for i in range(0, len(sheets)):
            # booksheet = workBook.get_sheet_by_name(sheets[i])
            booksheet = workBook[sheets[i]]
            print(booksheet.title)
            # 获取sheet页的行数据
            rows = booksheet.rows
            # 获取sheet页的列数据
            columns = booksheet.columns
            j = 0
            # 迭代所有的行
            for row in rows:
                j = j + 1
                line = [col.value for col in row]
                # print(j, line)

                # 获取每一行的数据
                cell_data_1 = booksheet.cell(row=j, column=1).value  # ERP
                cell_data_2 = booksheet.cell(row=j, column=2).value  # 生产商
                cell_data_3 = booksheet.cell(row=j, column=3).value  # 国家批文
                cell_data_4 = booksheet.cell(row=j, column=4).value  # 药品ID
                cell_data_5 = booksheet.cell(row=j, column=5).value  # 药品简称
                print(cell_data_1, cell_data_2, cell_data_3, cell_data_4, cell_data_5)
                medicine_values.append([cell_data_1, cell_data_2, cell_data_3, cell_data_4, cell_data_5])

        print("获取数据总条目：", len(medicine_values))
        medicine_values.pop(0)
        print(medicine_values)
        return medicine_values

        # 解析数据

    def parseHtml(self, data, keys):
        # print(data)
        soup = BeautifulSoup(data, "lxml")

        detail_urls = []
        img_urls = []
        img_list = soup.find_all("div", "fl h-drugs-pic bor")
        item_list = []
        if len(img_list) > 7:
            for i in range(0, 3):
                item_list.append(img_list[i])

        for div in item_list:
            try:
                # title = div.find("a").attrs["title"]
                detail_url = "http://yao.xywy.com" + div.find("a").attrs["href"]
                img_url = div.find("img").attrs["src"]
                img_name = div.find("img").attrs["title"]
            except Exception as e:
                print("Exception#", e)

            print(detail_url, img_url, img_name)

            detail_urls.append(detail_url)
            img_urls.append(img_url)

        print(len(detail_urls))

        index = 0
        # 过入详细页
        for url in detail_urls:
            print("请求详细：", url)
            content = self.requestUrl(url)
            # 找不了，就不在向下执行了
            if self.parseDetailHtml(img_urls[index], content, keys):
                break
            index += 1

    # 解析详细
    def parseDetailHtml(self, img_url, data, keys):
        # print(data)
        soup = BeautifulSoup(data, "html.parser")
        # soup = BeautifulSoup(data, "lxml")
        isFinding = False
        try:
            div = soup.find("div", "phonebox fl pz-box")
            dts = div.find("span", "r-rumbox").find_all("b")
            text = {}
            orders = []
            for dt in dts:
                label = dt.get_text()
                position = dt.attrs["style"]
                position = position[5:]
                position = position.replace("px;", "")
                position = position.strip()
                label = label.strip()
                text[position] = label
                orders.append(int(position))

            # 从小到大
            # orders.sort(reverse=True)
            orders.sort(reverse=False)
            length = len(orders)
            name = ""
            for i in range(0, length):
                name += text[str(orders[i])]
            text = name
            print(text, "-----------------批注文号#" + keys[2])

            # 判断两个ID是否相同
            key_word = keys[2]
            key_word = key_word.replace("国药准字", "")
            if text == key_word:
                isFinding = True
                print("列表图片地址：", img_url)
                path = "{0}/{1}".format(ROOT_PATH, keys[0])
                index = 0
                name = "{0}".format(index)
                # name = "{0}_{1}_{2}".format(keys[4], keys[1], index)
                self.download_image(img_url, path, name)
                lis = soup.find("div", "jqzoom_slideBox fl pr")
                imgs = lis.find_all("img")
                for img in imgs:
                    url = img.attrs["src"]
                    index += 1
                    name = "{0}".format(index)
                    self.download_image(url, path, name)
        except Exception as e:
            print("Exception#", e)
            logging.exception("#### 解析异常 #### " + keys[4])

        return isFinding

    def spider(self, source):
        data = self.get_resource(source)
        for key in data:
            print(key)
            logging.info(key)
            url_key = key[4]
            url_key = url_key.replace("(OTC)", "")
            url_key = url_key.replace("(处方药)", "")
            # 带公司名称的关键词
            # u_key = "{0} {1}".format(url_key, key[1][2:4])
            u_key = "{0}+{1}".format(url_key, key[1][0:4])
            url_key = u_key

            print("********************* 关键字：", url_key)
            logging.info("********************* 关键字：" + url_key)
            try:
                data = self.requestUrl(BASE_URL.format(url_key))
                self.parseHtml(data, key)
            except Exception as e:
                print(e)
                logging.exception("#### 找不到关键字 #### " + url_key)
                print("spider#### 找不到关键字 ####", url_key)

        self.close()

    # 下载图片
    def download_image(self, url, path, name):
        if not os.path.exists(path):
            os.makedirs(path)

        file = path + "/" + name + ".jpg"
        print("下载路径：", file, url)
        logging.info("#### 下载路径 #### " + file)
        logging.info("#### 下载URL #### " + url)

        out = open(file, "wb")
        out.write(requests.get(url).content)
        out.close()


def start_1():
    spider = MedicineSpider_YAOPIN()
    spider.spider(RESOURCE)


def start():
    # key = ["1616", "上海信谊黄河制药有限公司", "国药准字H31020147", "1251196", "叶酸片(处方药)"]
    # key = [25091, '上海雷允上药业有限公司', '国药准字Z31020515', 1212544, '牛黄解毒片(处方药)']
    key = [2916, '河北赛克药业有限公司', '国药准字H13023971', 2143666, '尼群地平片(处方药)']
    spider = MedicineSpider_YAOPIN()
    url_key = "尼群地平片+河北赛克"
    url = BASE_URL.format(url_key, url_key)
    data = spider.requestUrl(url)
    spider.parseHtml(data, key)
    spider.close()


if __name__ == "__main__":
    print("开始运行....")
    dic = {'-49': '1', '-35': '2', '-56': '3', '-63': 'H', '-7': '0', '-28': '1', '-42': '0', '-14': '2', '-21': '2'}
    # arr = [-63, -56, -49, -42, -35, -28, -21, -14, -7]
    arr = [-49, -35, -56, -63, -7, -28, -42, -14, -21]
    # 从小到大
    # arr.sort(reverse=True)
    # arr.sort(reverse=False)
    #
    # length = len(arr)
    # text = ""
    # for i in range(0, length):
    #     text += dic[str(arr[i])]
    # print(text)
    # print(arr)
    # start()
    start_1()
