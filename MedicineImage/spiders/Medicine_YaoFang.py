# -*- coding: utf-8 -*-
import scrapy
from MedicineImage.items import MedicineimageItem
from openpyxl import Workbook,load_workbook

def get_resource(resource):
    print(resource)
    # 找到需要xlsx文件的位置
    workBook = load_workbook(resource)

    # 如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
    # sheets = workBook.get_sheet_names()  # 从名称获取sheet
    sheets = workBook.sheetnames  # 从名称获取sheet
    print(sheets)

    medicine_values = []

    # 通过名字打开Sheet
    for i in range(0, len(sheets)):
        # booksheet = workBook.get_sheet_by_name(sheets[i])
        booksheet = workBook[sheets[i]]
        print(booksheet.title)
        # 获取sheet页的行数据
        rows = booksheet.rows
        # 获取sheet页的列数据
        columns = booksheet.columns
        j = 0
        # 迭代所有的行
        for row in rows:
            j = j + 1
            line = [col.value for col in row]
            # print(j, line)

            # 获取每一行的数据
            cell_data_1 = booksheet.cell(row=j, column=1).value  # 药品编号ID
            cell_data_2 = booksheet.cell(row=j, column=2).value  # 药品名称
            cell_data_3 = booksheet.cell(row=j, column=3).value  # 药品简称
            cell_data_4 = booksheet.cell(row=j, column=4).value  # 重量
            cell_data_5 = booksheet.cell(row=j, column=5).value  # 包装
            cell_data_6 = booksheet.cell(row=j, column=6).value  # 类型
            cell_data_7 = booksheet.cell(row=j, column=7).value  # 厂家
            cell_data_8 = booksheet.cell(row=j, column=8).value  # 日期
            print(cell_data_1, cell_data_2, cell_data_3, cell_data_4, cell_data_5, cell_data_6, cell_data_7,
                  cell_data_8)
            medicine_values.append(
                [cell_data_1, cell_data_2, cell_data_3, cell_data_4, cell_data_5, cell_data_6, cell_data_7,
                 cell_data_8])

    print("获取数据总条目：", len(medicine_values))
    medicine_values.pop(0)
    print(medicine_values)
    return medicine_values


class MedicineYaofangSpider(scrapy.Spider):
    name = 'Medicine-YaoFang'
    allowed_domains = ['yaofangwang.com']


    # keys = get_resource("./resource/商品.xlsx")
    # urls = []
    # base_url = "https://www.yaofangwang.com/search.html?keyword="
    # for i in range(0, len(keys)):
    #     url = base_url + keys[i][1] + keys[i][6]
    #     print(url)
    #     urls.append(url)
    #
    # print("请求链接的个数#", len(urls))
    # start_urls = urls
    start_urls = ["https://www.yaofangwang.com/search.html?keyword=布洛芬缓释胶囊(芬必得)中美天津史克制药有限公司"]

    header = {
        "User-Agent": "mozilla/5.0 (macintosh; intel mac os x 10_13_6) applewebkit/537.36 (khtml, like gecko) chrome/75.0.3770.142 safari/537.36"
    }

    def parse(self, response):
        # pass
        detail_list = response.xpath('//div[@class="info medicineInfo"]/a//@href').extract()

        picture_list = response.xpath('//img[@class="imgzoom"]//@src').extract()
        name_list = response.xpath('//img[@class="imgzoom"]//@alt').extract()
        # for pic in picture_list:
        for i in range(0, len(picture_list)):
            pic_url = "http:"+picture_list[i]
            name = name_list[i]
            print("----------->", pic_url)
            item = MedicineimageItem()
            item['url'] = [pic_url]
            item['name'] = [name]
            yield item

        for pic in detail_list:
            # url = "http:"+pic[2:]
            url = "http:"+pic
            print("----------->", url)
            yield scrapy.Request(url=url, callback=self.parse)