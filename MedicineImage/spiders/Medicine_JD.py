# -*- coding: utf-8 -*-
import scrapy
from MedicineImage.items import MedicineimageItem
from openpyxl import Workbook,load_workbook
from urllib.parse import quote,unquote

item_data = ["布洛芬缓释胶囊(芬必得)"]


def get_resource(resource):
    print(resource)
    # 找到需要xlsx文件的位置
    workBook = load_workbook(resource)

    # 如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
    # sheets = workBook.get_sheet_names()  # 从名称获取sheet
    sheets = workBook.sheetnames  # 从名称获取sheet
    print(sheets)

    medicine_values = []

    # 通过名字打开Sheet
    for i in range(0, len(sheets)):
        # booksheet = workBook.get_sheet_by_name(sheets[i])
        booksheet = workBook[sheets[i]]
        print(booksheet.title)
        # 获取sheet页的行数据
        rows = booksheet.rows
        # 获取sheet页的列数据
        columns = booksheet.columns
        j = 0
        # 迭代所有的行
        for row in rows:
            j = j + 1
            line = [col.value for col in row]
            # print(j, line)

            # 获取每一行的数据
            cell_data_1 = booksheet.cell(row=j, column=1).value  # 药品编号ID
            cell_data_2 = booksheet.cell(row=j, column=2).value  # 药品名称
            cell_data_3 = booksheet.cell(row=j, column=3).value  # 药品简称
            print(cell_data_1, cell_data_2, cell_data_3)
            medicine_values.append([cell_data_1, cell_data_2, cell_data_3])

    print("获取数据总条目：", len(medicine_values))
    medicine_values.pop(0)
    print(medicine_values)
    return medicine_values



class MedicineJdSpider(scrapy.Spider):
    name = 'Medicine-JD'
    allowed_domains = ['jd.com']
    start_urls = [
        'http://jd.com/',
        "https://search.jd.com/Search?keyword=%E6%A0%BC%E5%88%97%E7%BE%8E%E8%84%B2%E7%89%87(%E5%A4%84%E6%96%B9%E8%8D%AF)&enc=utf-8&wq=%E6%A0%BC%E5%88%97%E7%BE%8E%E8%84%B2%E7%89%87(%E5%A4%84%E6%96%B9%E8%8D%AF)&pvid=e594be63471b41b2a6821999b9c2e65a",
    ]

    # keys = get_resource("./resource/商品100.xlsx")
    # urls = []
    # base_url = "https://search.jd.com/Search?keyword={0}&enc=utf-8&wq={1}&pvid=e594be63471b41b2a6821999b9c2e65a"
    # for i in range(0, len(keys)):
    #     key = keys[i][2]
    #     # key = quote(key, "utf-8")
    #     url = base_url.format(key, key)
    #     print(url)
    #     urls.append(url)
    #
    # print("请求链接的个数#", len(urls))

    # start_urls = urls

    def parse(self, response):
        # pass

        # print(response.text)

        #  商品列表
        goods_list = response.xpath('//div[@class="goods-list-v2 gl-type-1 J-goods-list"]')
        # 商品详细
        # detail_list = goods_list.xpath('//div[@class="p-img"]/a//@href').extract()
        detail_list = goods_list.xpath('//div[@class="p-img"]/a//@href').extract_first()
        # 商品图片
        img_list = goods_list.xpath('//div[@class="p-img"]/a//img//@source-data-lazy-img').extract()

        # 列表图
        for url in img_list:
            img_url = "http:" +url
            print("图片：" + img_url)

            item = MedicineimageItem()
            item['url'] = [img_url]
            item['name'] = [item_data[0]]
            yield item

        # 解析详细页图片
        # zoom_img_list = response.xpath('//div[@class="jqzoom main-img"]/img').extract()
        zoom_img_list = response.xpath('//div[@class="jqzoom main-img"]/img//@data-origin').extract()
        zoom_img__alt_list = response.xpath('//div[@class="jqzoom main-img"]/img//@alt').extract()
        # zoom_img_zoom_list = response.xpath('//div[@class="layer-img"]/img//@data-origin').extract()
        # item_detail_list = response.xpath('//div[@class="Ptable-item"]/dl/dl')
        # for dl in item_detail_list:
        #     item_label = dl.xmpath("./dt/text()").extract()
        #     item_name = dl.xmpath("./dd/text()").extract()
        #     print("详细：" + item_label + ":"+item_name)

        for i in range(0, len(zoom_img_list)):
        # for url in zoom_img_list:
            img_url= "http:"+zoom_img_list[i]
            img_name= "http:"+zoom_img__alt_list[i]
            print("详细图片：" + img_url +" "+img_name)

            item = MedicineimageItem()
            item['url'] = [img_url]
            item['name'] = [item_data[0]]
            yield item


        if isinstance([], list):
            # 跳转到详细页面
            for url in detail_list:
                # 详细
                detail_url = "http:" + url
                print("----------->", detail_url)
                yield scrapy.Request(url=detail_url, callback=self.parse)
        else:
            detail_url = "http:" + detail_list
            print("----------->", detail_url)
            yield scrapy.Request(url=detail_url, callback=self.parse)



