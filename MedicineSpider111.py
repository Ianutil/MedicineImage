#!/usr/bin/python
# -*- coding: UTF-8 
# author: Ian
# Please,you must believe yourself who can do it beautifully !
"""
Are you OK?
https://www.111.com.cn
医药网  1药网

"""

from bs4 import BeautifulSoup
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import os
from openpyxl import Workbook, load_workbook
import logging

BASE_URL = "https://www.111.com.cn"
ROOT_PATH = "./file/image"
RESOURCE = r"./resource/商品100.xlsx"

search_js = """
$("#combobox-placeholder").attr("value","值1234");
$("#word").attr("value","值1234");

function searchBtn() {　
    search();
};
searchBtn()

"""


class MedicineSpider111(object):
    keys = []

    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # 使用无头谷歌浏览器模式
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        # 指定谷歌浏览器路径
        # webdriver.Chrome(chrome_options=chrome_options, executable_path='/usr/local/bin/chromedriver')
        webdriver.Chrome(options=chrome_options)
        self.driver = webdriver.Chrome()

        file_name = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        file = "./resource/{0}.log".format(file_name)
        logging.basicConfig(filename=file, filemode="w", format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                            datefmt="%d-%M-%Y %H:%M:%S", level=logging.DEBUG)

    def close(self):
        self.driver.close()
        self.driver.quit()

    def requestUrl(self, url):
        self.driver.get(url=url)
        # 等待1s
        time.sleep(1)

        # 获取响应内容
        content = self.driver.page_source.encode('utf-8')
        content = str(content, "utf-8")
        return content

    def get_resource(self, resource):
        print(resource)
        # 找到需要xlsx文件的位置
        workBook = load_workbook(resource)

        # 如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
        # sheets = workBook.get_sheet_names()  # 从名称获取sheet
        sheets = workBook.sheetnames  # 从名称获取sheet
        print(sheets)

        medicine_values = []

        # 通过名字打开Sheet
        for i in range(0, len(sheets)):
            # booksheet = workBook.get_sheet_by_name(sheets[i])
            booksheet = workBook[sheets[i]]
            print(booksheet.title)
            # 获取sheet页的行数据
            rows = booksheet.rows
            # 获取sheet页的列数据
            columns = booksheet.columns
            j = 0
            # 迭代所有的行
            for row in rows:
                j = j + 1
                line = [col.value for col in row]
                # print(j, line)

                # 获取每一行的数据
                cell_data_1 = booksheet.cell(row=j, column=1).value  # ERP
                cell_data_2 = booksheet.cell(row=j, column=2).value  # 生产商
                cell_data_3 = booksheet.cell(row=j, column=3).value  # 国家批文
                cell_data_4 = booksheet.cell(row=j, column=4).value  # 药品ID
                cell_data_5 = booksheet.cell(row=j, column=5).value  # 药品简称
                print(cell_data_1, cell_data_2, cell_data_3, cell_data_4, cell_data_5)
                medicine_values.append([cell_data_1, cell_data_2, cell_data_3, cell_data_4, cell_data_5])

        print("获取数据总条目：", len(medicine_values))
        medicine_values.pop(0)
        print(medicine_values)
        return medicine_values

    def searchKey(self, key):
        key = search_js.replace("值1234", key)
        # 执行js滚动浏览器窗口到底部
        self.driver.execute_script(key)
        time.sleep(1)
        content = self.driver.page_source.encode('utf-8')
        content = str(content, "utf-8")
        return content

    # 解析搜索列表页
    def parseHtml(self, url_key, key):
        # 搜索结果
        content = self.searchKey(url_key)
        soup = BeautifulSoup(content, "html.parser")
        item_divs = soup.find_all("div", "itemSearchResultCon")
        # 设置范围
        if len(item_divs) > 7:
            item_divs = item_divs[:6]
        for div in item_divs:
            # print("----------->", div)
            if div == '\n':
                continue
            logging.info(div)
            try:
                img_url = div.find("img").attrs["src"]
                detail_url = "http:" +div.find("a").attrs["href"]
                # detail_url = "http:" + div.attrs["href"]
                logging.info(img_url)
                logging.info(detail_url)
                print(img_url, detail_url)
                flag = self.spiderDetail(img_url, detail_url, key)
                if flag:
                    return True
            except Exception as e:
                print(e)
        return False

    def spider(self, source):
        self.requestUrl(BASE_URL)
        data = self.get_resource(source)
        for key in data:
            print(key)
            logging.info(key)
            url_key = key[4]
            url_key = url_key.replace("(OTC)", "")
            url_key = url_key.replace("(处方药)", "")
            # 带公司名称的关键词
            u_key = "{0} {1}".format(url_key, key[1][2:4])
            url_key = u_key

            print("********************* 关键字：", url_key)
            logging.info("********************* 关键字：" + url_key)

            if not self.parseHtml(url_key, key):
                print("末找到关键字：", url_key)
                logging.info("末找到关键字：" + url_key)

        self.close()

    # 下载图片
    def download_image(self, url, path, name):
        if not os.path.exists(path):
            os.makedirs(path)

        file = path + "/" + name + ".jpg"
        print("下载路径：", file, url)
        logging.info("#### 下载路径 #### " + file)
        logging.info("#### 下载URL #### " + url)

        out = open(file, "wb")
        out.write(requests.get(url).content)
        out.close()

    # 解析详细
    def spiderDetail(self, img_url, detail_url, keys):
        data = self.requestUrl(detail_url)
        # print(data)
        soup = BeautifulSoup(data, "html.parser")
        div = soup.find("div", "goods_intro")
        # print(div)
        # print("**********************")

        logging.info(div)
        logging.info("**********************")

        ths = div.find_all("th")
        tds = div.find_all("td")
        logging.info(ths)
        logging.info("**********************")
        logging.info(tds)

        # print(ths)
        # print("**********************")
        # print(tds)
        for i in range(0, len(tds)):
            th = ths[i].get_text().strip()
            td = tds[i].get_text().strip()
            if th == "批准文号：":
                td = td.replace("(国家食药局查询)", "")
            print(th, td)
            logging.info(th+""+td)
            # 判断是否为需要查询的批号
            if th == "批准文号：" and td == keys[2]:
                path = "{0}/{1}".format(ROOT_PATH, keys[0])
                index = 0
                name = "{0}".format(index)
                self.download_image(img_url, path, name)

                # 图片列表
                div = soup.find("div", "dt_scrollable")
                imgs = div.find_all("a")
                try:
                    for img in imgs:
                        # print(img)
                        value = img.attrs["rel"]
                        # print("******************")
                        print(value)
                        # print(len(value))
                        img_url_detail = value[3]
                        img_url_detail = img_url_detail.replace("}", "")
                        img_url_detail = img_url_detail.replace("\'", "")
                        img_url_detail = img_url_detail[11:]
                        print(img_url_detail)
                        index += 1
                        name = "{0}".format(index)
                        self.download_image(img_url_detail, path, name)
                        return True
                except Exception as e:
                    print("spiderDetail", e)

        return False

def start_1():
    spider = MedicineSpider111()
    source = "./resource/medicine.xlsx"
    spider.spider(source)


def start():
    key = ["289202", "江苏晨牌邦德药业有限公司", "国药准字Z51020212", "1251196", "伤湿止痛膏(OTC)"]
    spider = MedicineSpider111()
    spider.requestUrl(BASE_URL)
    spider.parseHtml("复方黄连素片 太极", key)
    spider.close()


if __name__ == "__main__":
    print("开始运行....")
    # text = "国药准字Z20055103(国家食药局查询)"
    # tx = text.replace("(国家食药局查询)", "")
    # print(tx)

    # start()
    start_1()